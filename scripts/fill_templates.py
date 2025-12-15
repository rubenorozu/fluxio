#!/usr/bin/env python3
"""
Script para llenar las plantillas Excel con datos de ejemplo
"""
import openpyxl
from pathlib import Path

# Rutas de las plantillas
base_path = Path(__file__).parent.parent / "1 recursos" / "Plantillas"
plantilla_espacios = base_path / "plantilla_espacios.xlsx"
plantilla_equipos = base_path / "plantilla_equipos.xlsx"
plantilla_talleres = base_path / "plantilla_talleres.xlsx"

# Datos para ESPACIOS (10 registros)
# Columnas: nombre, descripcion, estado, responsable_email, requiere_reserva_espacio, tiempo_anticipacion
espacios_data = [
    {
        "nombre": "Auditorio Principal",
        "descripcion": "Auditorio con capacidad para 500 personas, equipado con sistema de audio profesional",
        "estado": "DISPONIBLE",
        "responsable_email": "auditorio@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 48
    },
    {
        "nombre": "Oficina Administrativa",
        "descripcion": "Oficina equipada con escritorios, computadoras y mobiliario de oficina",
        "estado": "DISPONIBLE",
        "responsable_email": "admin@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 24
    },
    {
        "nombre": "Polideportivo",
        "descripcion": "Instalación deportiva multiusos con cancha de basketball, volleyball y futbol rápido",
        "estado": "DISPONIBLE",
        "responsable_email": "deportes@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 72
    },
    {
        "nombre": "Sala de Juntas Ejecutiva",
        "descripcion": "Sala de reuniones con mesa para 20 personas, proyector y videoconferencia",
        "estado": "DISPONIBLE",
        "responsable_email": "juntas@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 12
    },
    {
        "nombre": "Salón de Clases 101",
        "descripcion": "Aula con capacidad para 40 estudiantes, pizarrón inteligente y proyector",
        "estado": "DISPONIBLE",
        "responsable_email": "academico@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 24
    },
    {
        "nombre": "Salón de Usos Múltiples",
        "descripcion": "Espacio versátil para eventos, talleres y conferencias",
        "estado": "DISPONIBLE",
        "responsable_email": "eventos@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 48
    },
    {
        "nombre": "Laboratorio de Cómputo",
        "descripcion": "Laboratorio equipado con 30 computadoras de última generación",
        "estado": "DISPONIBLE",
        "responsable_email": "computo@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 24
    },
    {
        "nombre": "Biblioteca Central",
        "descripcion": "Espacio de estudio con área de lectura, cubículos individuales y sala de estudio grupal",
        "estado": "DISPONIBLE",
        "responsable_email": "biblioteca@institucion.edu",
        "requiere_reserva_espacio": "NO",
        "tiempo_anticipacion": 0
    },
    {
        "nombre": "Cafetería Institucional",
        "descripcion": "Área de comedor con mesas para 150 personas",
        "estado": "DISPONIBLE",
        "responsable_email": "cafeteria@institucion.edu",
        "requiere_reserva_espacio": "NO",
        "tiempo_anticipacion": 0
    },
    {
        "nombre": "Sala de Conferencias",
        "descripcion": "Sala equipada para videoconferencias y presentaciones ejecutivas",
        "estado": "DISPONIBLE",
        "responsable_email": "conferencias@institucion.edu",
        "requiere_reserva_espacio": "SI",
        "tiempo_anticipacion": 24
    }
]

# Datos para EQUIPOS (10 registros)
# Columnas: nombre, descripcion, numero_serie, activo_fijo, estado, espacio_asignado, fijo_a_espacio, responsable_email, tiempo_anticipacion
equipos_data = [
    {
        "nombre": "Cámara Mirrorless Sony A7III",
        "descripcion": "Cámara profesional de fotograma completo con lente 24-70mm",
        "numero_serie": "CAM-SONY-001",
        "activo_fijo": "AF-2024-001",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 24
    },
    {
        "nombre": "Cámara de TV Broadcast",
        "descripcion": "Cámara profesional para producción de televisión con trípode",
        "numero_serie": "CAM-BROAD-001",
        "activo_fijo": "AF-2024-002",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 48
    },
    {
        "nombre": "Grabadora de Campo Zoom H6",
        "descripcion": "Grabadora de audio portátil de 6 canales",
        "numero_serie": "GRAB-ZOOM-001",
        "activo_fijo": "AF-2024-003",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 12
    },
    {
        "nombre": "Laptop Dell XPS 15",
        "descripcion": "Laptop de alto rendimiento con procesador i7, 16GB RAM, 512GB SSD",
        "numero_serie": "LAP-DELL-001",
        "activo_fijo": "AF-2024-004",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "computo@institucion.edu",
        "tiempo_anticipacion": 24
    },
    {
        "nombre": "Micrófono Lavalier Inalámbrico",
        "descripcion": "Sistema de micrófono inalámbrico de solapa para entrevistas",
        "numero_serie": "MIC-LAV-001",
        "activo_fijo": "AF-2024-005",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 12
    },
    {
        "nombre": "Micrófono Shure SM58",
        "descripcion": "Micrófono dinámico vocal profesional",
        "numero_serie": "MIC-SHURE-001",
        "activo_fijo": "AF-2024-006",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 12
    },
    {
        "nombre": "Cable XLR 5 metros",
        "descripcion": "Cable balanceado profesional para audio",
        "numero_serie": "CABLE-XLR-001",
        "activo_fijo": "AF-2024-007",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 0
    },
    {
        "nombre": "Cable Plug TRS 3 metros",
        "descripcion": "Cable de audio estéreo 1/4 pulgadas",
        "numero_serie": "CABLE-PLUG-001",
        "activo_fijo": "AF-2024-008",
        "estado": "DISPONIBLE",
        "espacio_asignado": "",
        "fijo_a_espacio": "NO",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 0
    },
    {
        "nombre": "Proyector Epson 5000 Lúmenes",
        "descripcion": "Proyector de alta luminosidad Full HD",
        "numero_serie": "PROY-EPSON-001",
        "activo_fijo": "AF-2024-009",
        "estado": "DISPONIBLE",
        "espacio_asignado": "Auditorio Principal",
        "fijo_a_espacio": "SI",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 24
    },
    {
        "nombre": "Consola Mezcladora Behringer X32",
        "descripcion": "Consola digital de 32 canales para audio profesional",
        "numero_serie": "CONS-BEH-001",
        "activo_fijo": "AF-2024-010",
        "estado": "DISPONIBLE",
        "espacio_asignado": "Auditorio Principal",
        "fijo_a_espacio": "SI",
        "responsable_email": "audiovisual@institucion.edu",
        "tiempo_anticipacion": 48
    }
]

# Datos para TALLERES (10 registros)
# Columnas: nombre, descripcion, capacidad, profesor, horario, salon, fecha_inicio, fecha_fin, inscripciones_abiertas, responsable_email
talleres_data = [
    {
        "nombre": "Inducción Institucional",
        "descripcion": "Taller de bienvenida para nuevos integrantes de la organización",
        "capacidad": 30,
        "profesor": "Recursos Humanos",
        "horario": "Lunes y Miércoles 9:00-11:00",
        "salon": "Sala de Conferencias",
        "fecha_inicio": "2025-01-15",
        "fecha_fin": "2025-01-29",
        "inscripciones_abiertas": "SI",
        "responsable_email": "rh@institucion.edu"
    },
    {
        "nombre": "Capacitación en Excel Avanzado",
        "descripcion": "Curso de Excel con tablas dinámicas, macros y análisis de datos",
        "capacidad": 25,
        "profesor": "Lic. María González",
        "horario": "Martes y Jueves 14:00-18:00",
        "salon": "Laboratorio de Cómputo",
        "fecha_inicio": "2025-01-20",
        "fecha_fin": "2025-02-10",
        "inscripciones_abiertas": "SI",
        "responsable_email": "capacitacion@institucion.edu"
    },
    {
        "nombre": "Manejo de Cámaras de Video",
        "descripcion": "Taller práctico de operación de cámaras profesionales y composición",
        "capacidad": 15,
        "profesor": "Mtro. Carlos Ramírez",
        "horario": "Viernes 10:00-16:00",
        "salon": "Salón de Usos Múltiples",
        "fecha_inicio": "2025-01-22",
        "fecha_fin": "2025-02-12",
        "inscripciones_abiertas": "SI",
        "responsable_email": "audiovisual@institucion.edu"
    },
    {
        "nombre": "Administración del Tiempo",
        "descripcion": "Técnicas y herramientas para optimizar la gestión del tiempo laboral",
        "capacidad": 40,
        "profesor": "Dra. Ana Martínez",
        "horario": "Miércoles 16:00-18:00",
        "salon": "Auditorio Principal",
        "fecha_inicio": "2025-01-25",
        "fecha_fin": "2025-02-15",
        "inscripciones_abiertas": "SI",
        "responsable_email": "desarrollo@institucion.edu"
    },
    {
        "nombre": "Optimización de Recursos",
        "descripcion": "Estrategias para maximizar el uso eficiente de recursos institucionales",
        "capacidad": 35,
        "profesor": "Ing. Roberto López",
        "horario": "Lunes 15:00-18:00",
        "salon": "Sala de Juntas Ejecutiva",
        "fecha_inicio": "2025-02-01",
        "fecha_fin": "2025-02-22",
        "inscripciones_abiertas": "SI",
        "responsable_email": "administracion@institucion.edu"
    },
    {
        "nombre": "Manejo de Editores de Texto",
        "descripcion": "Curso de Word, Google Docs y herramientas de edición profesional",
        "capacidad": 30,
        "profesor": "Lic. Patricia Hernández",
        "horario": "Martes y Jueves 10:00-12:00",
        "salon": "Laboratorio de Cómputo",
        "fecha_inicio": "2025-02-05",
        "fecha_fin": "2025-02-25",
        "inscripciones_abiertas": "SI",
        "responsable_email": "capacitacion@institucion.edu"
    },
    {
        "nombre": "Uso Adecuado de Equipos Técnicos",
        "descripcion": "Capacitación en el manejo y cuidado de equipos audiovisuales",
        "capacidad": 20,
        "profesor": "Téc. Juan Pérez",
        "horario": "Viernes 14:00-16:00",
        "salon": "Salón de Usos Múltiples",
        "fecha_inicio": "2025-02-08",
        "fecha_fin": "2025-02-28",
        "inscripciones_abiertas": "SI",
        "responsable_email": "audiovisual@institucion.edu"
    },
    {
        "nombre": "Liderazgo y Trabajo en Equipo",
        "descripcion": "Desarrollo de habilidades de liderazgo y colaboración efectiva",
        "capacidad": 30,
        "profesor": "Coach Laura Sánchez",
        "horario": "Miércoles y Viernes 9:00-13:00",
        "salon": "Sala de Conferencias",
        "fecha_inicio": "2025-02-10",
        "fecha_fin": "2025-03-05",
        "inscripciones_abiertas": "SI",
        "responsable_email": "desarrollo@institucion.edu"
    },
    {
        "nombre": "Seguridad e Higiene Laboral",
        "descripcion": "Normativas y prácticas de seguridad en el entorno de trabajo",
        "capacidad": 50,
        "profesor": "Ing. Miguel Torres",
        "horario": "Jueves 11:00-13:00",
        "salon": "Auditorio Principal",
        "fecha_inicio": "2025-02-15",
        "fecha_fin": "2025-03-01",
        "inscripciones_abiertas": "SI",
        "responsable_email": "seguridad@institucion.edu"
    },
    {
        "nombre": "Comunicación Efectiva",
        "descripcion": "Técnicas de comunicación oral y escrita para el ámbito profesional",
        "capacidad": 35,
        "profesor": "Lic. Sofía Morales",
        "horario": "Lunes y Miércoles 13:00-16:00",
        "salon": "Salón de Clases 101",
        "fecha_inicio": "2025-02-18",
        "fecha_fin": "2025-03-10",
        "inscripciones_abiertas": "SI",
        "responsable_email": "comunicacion@institucion.edu"
    }
]

def fill_espacios():
    """Llena la plantilla de espacios"""
    print("📍 Llenando plantilla de espacios...")
    wb = openpyxl.load_workbook(plantilla_espacios)
    ws = wb.active
    
    # Verificar encabezados
    headers = [cell.value for cell in ws[1]]
    print(f"   Columnas encontradas: {headers}")
    
    # Limpiar datos existentes (excepto encabezados)
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
    
    # Agregar datos
    # Columnas: nombre, descripcion, estado, responsable_email, requiere_reserva_espacio, tiempo_anticipacion
    for idx, espacio in enumerate(espacios_data, start=2):
        ws[f'A{idx}'] = espacio['nombre']
        ws[f'B{idx}'] = espacio['descripcion']
        ws[f'C{idx}'] = espacio['estado']
        ws[f'D{idx}'] = espacio['responsable_email']
        ws[f'E{idx}'] = espacio['requiere_reserva_espacio']
        ws[f'F{idx}'] = espacio['tiempo_anticipacion']
    
    wb.save(plantilla_espacios)
    print(f"   ✅ {len(espacios_data)} espacios agregados")

def fill_equipos():
    """Llena la plantilla de equipos"""
    print("🎥 Llenando plantilla de equipos...")
    wb = openpyxl.load_workbook(plantilla_equipos)
    ws = wb.active
    
    # Verificar encabezados
    headers = [cell.value for cell in ws[1]]
    print(f"   Columnas encontradas: {headers}")
    
    # Limpiar datos existentes (excepto encabezados)
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
    
    # Agregar datos
    # Columnas: nombre, descripcion, numero_serie, activo_fijo, estado, espacio_asignado, fijo_a_espacio, responsable_email, tiempo_anticipacion
    for idx, equipo in enumerate(equipos_data, start=2):
        ws[f'A{idx}'] = equipo['nombre']
        ws[f'B{idx}'] = equipo['descripcion']
        ws[f'C{idx}'] = equipo['numero_serie']
        ws[f'D{idx}'] = equipo['activo_fijo']
        ws[f'E{idx}'] = equipo['estado']
        ws[f'F{idx}'] = equipo['espacio_asignado']
        ws[f'G{idx}'] = equipo['fijo_a_espacio']
        ws[f'H{idx}'] = equipo['responsable_email']
        ws[f'I{idx}'] = equipo['tiempo_anticipacion']
    
    wb.save(plantilla_equipos)
    print(f"   ✅ {len(equipos_data)} equipos agregados")

def fill_talleres():
    """Llena la plantilla de talleres"""
    print("🎓 Llenando plantilla de talleres...")
    wb = openpyxl.load_workbook(plantilla_talleres)
    ws = wb.active
    
    # Verificar encabezados
    headers = [cell.value for cell in ws[1]]
    print(f"   Columnas encontradas: {headers}")
    
    # Limpiar datos existentes (excepto encabezados)
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
    
    # Agregar datos
    # Columnas: nombre, descripcion, capacidad, profesor, horario, salon, fecha_inicio, fecha_fin, inscripciones_abiertas, responsable_email
    for idx, taller in enumerate(talleres_data, start=2):
        ws[f'A{idx}'] = taller['nombre']
        ws[f'B{idx}'] = taller['descripcion']
        ws[f'C{idx}'] = taller['capacidad']
        ws[f'D{idx}'] = taller['profesor']
        ws[f'E{idx}'] = taller['horario']
        ws[f'F{idx}'] = taller['salon']
        ws[f'G{idx}'] = taller['fecha_inicio']
        ws[f'H{idx}'] = taller['fecha_fin']
        ws[f'I{idx}'] = taller['inscripciones_abiertas']
        ws[f'J{idx}'] = taller['responsable_email']
    
    wb.save(plantilla_talleres)
    print(f"   ✅ {len(talleres_data)} talleres agregados")

if __name__ == "__main__":
    print("\n🚀 Iniciando llenado de plantillas Excel...\n")
    
    try:
        fill_espacios()
        fill_equipos()
        fill_talleres()
        
        print("\n✨ ¡Todas las plantillas han sido actualizadas exitosamente!")
        print(f"\n📁 Ubicación: {base_path}")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
